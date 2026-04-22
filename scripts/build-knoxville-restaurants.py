#!/usr/bin/env python3
"""
Rebuild assets/data/knoxville-restaurants.json from a Knoxville allergy spreadsheet
(e.g. Knoxville_Allergy_Master_Expanded.xlsx).

Usage:
  python3 scripts/build-knoxville-restaurants.py /path/to/Knoxville_Allergy_Master_Expanded.xlsx

Coordinates are merged from COORDS_BY_NAME (update when venues move or after geocoding).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

def _norm_apostrophe(s: str) -> str:
    return s.replace("\u2019", "'").replace("\u2018", "'")


# Lat/lon from manual verification or geocoding; keys use straight apostrophes (matched after normalizing xlsx).
COORDS_BY_NAME: dict[str, tuple[float, float]] = {
    "Farmacy": (35.93913, -83.98736),
    "Yassin's Falafel House": (35.96412, -83.91985),
    "Good Golly Tamale": (35.9512, -83.9089),
    "The Plaid Apron": (35.9649, -83.9071),
    "Tomato Head": (35.96125, -83.91875),
    "Kefi": (35.9634, -83.9298),
    "Chivo Taqueria": (35.9601, -83.9169),
    "Sunspot": (35.9568, -83.9385),
    "Aretha Frankenstein's": (35.9732, -83.9236),
    "A Dopo": (35.9642, -83.9068),
    "Tupelo Honey": (35.96105, -83.91845),
    "Scrambled Jake's": (35.942, -83.988),
    "Wild Love Bakehouse": (35.985421, -83.934451),
    "Ruby Sunshine": (35.9613, -83.9185),
    "Mimosas": (35.9632, -83.9174),
    "French Market Creperie": (35.9608, -83.9185),
    "Not Watson's": (35.974, -83.919),
    "Treetop Coffee Shop": (35.9685, -83.926),
    "Intrepid Nitro Coffee": (35.958, -83.91),
    "K Brew": (35.983716, -83.9221265),
    "Honeybee Coffee Co.": (35.9655, -83.9205),
    "Awaken Coffee": (35.9662, -83.924),
    "Status Dough Coffee": (35.9602, -83.9188),
    "Remedy Coffee": (35.974536, -83.924611),
    "Five Thirty Coffee Roasters": (35.964, -83.925),
}

COLS = [
    "Gluten-Free",
    "Dairy-Free",
    "Egg-Free",
    "Vegetarian",
    "Peanut-Free",
    "Tree Nut-Free",
    "Soy-Free",
    "Shellfish-Free",
]


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    out_path = root / "assets" / "data" / "knoxville-restaurants.json"
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if not xlsx or not xlsx.is_file():
        print("Usage: build-knoxville-restaurants.py <Knoxville_Allergy_Restaurants.xlsx>", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    name_i = header.index("Restaurant")
    notes_i = header.index("Notes") if "Notes" in header else None
    what_i = header.index("What to Order") if "What to Order" in header else None

    restaurants = []
    for row in rows[1:]:
        if not row or row[name_i] is None:
            continue
        name = str(row[name_i]).strip()
        if not name:
            continue
        acc = {}
        for c in COLS:
            if c not in header:
                continue
            i = header.index(c)
            v = row[i] if i < len(row) else None
            acc[c] = str(v).strip() if v is not None else ""

        notes = ""
        if notes_i is not None and notes_i < len(row) and row[notes_i] is not None:
            notes = str(row[notes_i]).strip()

        what_to_order = ""
        if what_i is not None and what_i < len(row) and row[what_i] is not None:
            what_to_order = str(row[what_i]).strip()

        lookup = _norm_apostrophe(name)
        coords = next(
            (v for k, v in COORDS_BY_NAME.items() if _norm_apostrophe(k) == lookup),
            None,
        )
        if coords is None:
            print(f"Warning: add COORDS_BY_NAME entry for {name!r}", file=sys.stderr)
            continue
        lat, lon = coords

        restaurants.append(
            {
                "name": name,
                "notes": notes,
                "what_to_order": what_to_order,
                "lat": lat,
                "lon": lon,
                "accommodations": acc,
            }
        )

    payload = {"source": xlsx.name, "restaurants": restaurants}
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {len(restaurants)} restaurants to {out_path}")


if __name__ == "__main__":
    main()
